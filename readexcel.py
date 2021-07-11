from openpyxl import load_workbook
import pandas as pd
import requests
import pprint

excel_file = 'sudocontrol-original.xlsx' 
wb = load_workbook(excel_file, data_only = True)
sh = wb['Abraçadeira (2)']
wb.sheetnames
sh['A1'].fill.start_color.index



color_list = []
for i in range(7, sh.max_row):
    color_index = sh['A'+str(i)].fill.start_color.index
    if color_index != '00000000':
        color_list.append(color_index)


print(color_list)
numrows = len(color_list)
wbpd = pd.read_excel(excel_file, skiprows=5, sheet_name='Abraçadeira (2)', converters={'Pregão':str,'Cód. Unidade':str})
wbpd = wbpd[0:numrows]
color_col = [1 if i == 9 else 0 for i in color_list]
wbpd['Ativo'] = color_col

wbpd["N_PREGAO"] = wbpd['Cód. Unidade'] + "000" + wbpd['Pregão']

quant_item = []

for index, row in wbpd.iterrows():
    if row['Ativo'] == 1:
        num_item = int(row['Item'])
        try:
            response = requests.get('http://compras.dados.gov.br/pregoes/doc/pregao/{}/itens.json'.format(row["N_PREGAO"]))
            resdict = response.json()

            itens = resdict['_embedded']['pregoes']

            for item in itens:
                item_title = item['_links']['self']['title']
                if int(item_title.split()[1].split(":")[0]) == num_item:
                    quant = int(item['quantidade_item'])

            quant_item.append(quant)
            quant = 0      
        except:
            response = requests.get('http://compras.dados.gov.br/pregoes/doc/pregao/{}/itens.json?offset=500'.format(row["N_PREGAO"]))
            resdict = response.json()

            itens = resdict['_embedded']['pregoes']

            for item in itens:
                item_title = item['_links']['self']['title']
                if int(item_title.split()[1].split(":")[0]) == num_item:
                    quant = int(item['quantidade_item'])
            
            quant_item.append(quant)

    else:
        quant_item.append('Item não ativo')
        

len(quant_item)

response = requests.get('http://compras.dados.gov.br/pregoes/doc/pregao/1350260001312020/itens.json')
resdict = response.json()

itens = resdict['_embedded']['pregoes']

for item in itens:
    item_title = item['_links']['self']['title']
    if int(item_title.split()[1].split(":")[0]) == 803:
        quant = item['quantidade_item']

quant

pdf_links = []
for index, row in wbpd.iterrows():
    pdf_links.append("https://comprasnet.gov.br/livre/pregao/anexosDosItens.asp?uasg={}&numprp={}&prgcod=863000".format(row['Cód. Unidade'], row['Pregão']))

pdf_links
