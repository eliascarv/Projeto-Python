from openpyxl import load_workbook
from datetime import date
from statistics import mean, stdev, median
import pandas as pd

excel_file = '2021.05.13 - Relacao_de_itens_e_precos2.xlsx'
wb = load_workbook(excel_file , data_only = True)
sh = wb['Abraçadeira']
sh['A1404'].fill.start_color.index
abas = wb.sheetnames

resultado = pd.ExcelWriter('resultado_consultas_final.xlsx')

relatorio_lista = []

for aba in abas[1:len(abas)]:
    sh = wb[aba]
    color_list = []
    for i in range(7, sh.max_row):
        color_index = sh['A' + str(i)].fill.start_color.index
        if color_index != '00000000':
            color_list.append(color_index)


    numrows = len(color_list)
    wbpd = pd.read_excel(excel_file, skiprows = 5, sheet_name = aba, converters = {'Identif Compra': str,'Pregão':str,'Cód. Unidade':str})
    wbpd = wbpd[0:numrows]
    color_col = [1 if color == 'FFDAF2F4' else 0 for color in color_list]
    wbpd['Ativo'] = color_col

    wbpd["Número do Pregão"] = wbpd['Cód. Unidade'] + "000" + wbpd['Pregão']

    anexos_links = []

    for index, row in wbpd.iterrows():
        anexos_links.append("https://comprasnet.gov.br/livre/pregao/anexosDosItens.asp?uasg={}&numprp={}&prgcod=863000".format(row['Cód. Unidade'], row['Pregão']))

    wbpd['Anexos Pregão'] = anexos_links

    meses = {
        "Jan": 1, "Fev": 2, "Mar": 3, "Abr":  4, "Mai":  5, "Jun":  6, 
        "Jul": 7, "Ago": 8, "Set": 9, "Out": 10, "Nov": 11, "Dez": 12
    }

    mes_ano = wbpd['Mês Resultado Compra']

    mes = [meses[i.split()[0]] for i in mes_ano]
    ano = [int(i.split()[1]) for i in mes_ano]

    wbpd['Mês'] = mes
    wbpd['Ano'] = ano

    mes_atual = date.today().month
    ano_atual = date.today().year

    dentro_do_periodo = []
    for index, row in wbpd.iterrows():
        if row['Ativo'] == 1:
            if row['Ano'] == (ano_atual - 1) and row['Mês'] in range(mes_atual, 13):
                dentro_do_periodo.append(1)
            elif row['Ano'] == ano_atual and row['Mês'] in range(1, mes_atual + 1):
                dentro_do_periodo.append(1)
            else:
                dentro_do_periodo.append(0)
        else:
            dentro_do_periodo.append('Item não ativo')

    len(dentro_do_periodo)

    wbpd['Dentro do Período'] = dentro_do_periodo

    valores = []
    for index, row in wbpd.iterrows():
        if row['Ativo'] == 1 and row['Dentro do Período'] == 1:
            valores.append(row['Valor'])

    media = mean(valores)
    desvio = stdev(valores)
    mediana = median(valores)
    coeficiente = desvio / media
    preco = mediana if coeficiente > 0.25 else media

    relatorio_lista.append([aba, media, desvio, coeficiente, mediana, preco])

    wbpd.to_excel(resultado, aba, index = False)

relatorio = pd.DataFrame(
    relatorio_lista, 
    columns = ['Item', 'Média', 'Desvio Padão', 'Coeficiente', 'Mediana', 'Preço']
)

relatorio.to_excel(resultado, 'Relatorio', index = False)

resultado.save()